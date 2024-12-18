import requests
from bs4 import BeautifulSoup
from tabulate import tabulate
import openpyxl
from openpyxl.styles import PatternFill

login_url = "https://pfm.smartcitylk.org/wp-login.php"
target_url = "https://pfm.smartcitylk.org/wp-admin/admin.php?page=generalInfo"
username = "admin"
password = "Xfiles3384@"

session = requests.Session()

login_data = {
    'log': username,
    'pwd': password,
    'wp-submit': 'Log In',
    'redirect_to': target_url,
}
login_response = session.post(login_url, data=login_data)

if 'wp-admin' in login_response.url:
    target_response = session.get(target_url)

    if target_response.status_code == 200:
        soup = BeautifulSoup(target_response.content, 'html.parser')
        table = soup.find('table', {'id': 'table26'})

        x, y, z = 'Local Authority', 'Other Key', 'Another Key'
        local_authority_info = {cols[0].text.strip(): cols[1].text.strip() for cols in [row.find_all('td') for row in soup.find('table').find_all('tr') if len(row.find_all('td')) == 2]}
        N = local_authority_info.get(x, f"{x} not found")

        data = []

        for row in table.find_all('tr')[1:]:
            columns = row.find_all('td')

            if len(columns) >= 4:
                vehicle_type = columns[0].text.strip()

                input_2 = columns[2].find('input')
                input_3 = columns[3].find('input')

                if input_2 and input_3 and 'value' in input_2.attrs and 'value' in input_3.attrs:
                    in_running_condition = int(input_2['value']) if input_2['value'].isdigit() else None
                    not_running_condition = int(input_3['value']) if input_3['value'].isdigit() else None

                    data.append({
                        'local_authority': N,
                        'Vehicle Type': vehicle_type,
                        'In Running Condition': in_running_condition,
                        'Not Running Condition': not_running_condition
                    })

        if data:
            headers = data[0].keys()
            table_data = [[entry[col] for col in headers] for entry in data]

            wb = openpyxl.Workbook()
            ws = wb.active

            header_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

            ws.append(list(headers))
            for cell in ws[1]:
                cell.fill = header_fill

            for row in table_data:
                ws.append(row)

            excel_filename = f"{N}.xlsx"
            wb.save(excel_filename)
            print(f" {excel_filename}")
        else:
            print("No data to display.")
    else:
        print(f"Failed to retrieve the target webpage. Status code: {target_response.status_code}")
else:
    print("Login failed. Please check your credentials.")
    